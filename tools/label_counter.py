"""
标签出现次数统计工具

功能说明：
- 递归扫描选定目录中的所有 Labelme JSON 文件
- 按用户上传的标签列表顺序，统计每个 JSON 文件内各标签的出现次数
- 生成 xlsx 格式统计报告，未出现的标签显示为 0
"""

import csv
import json
import os

from openpyxl import Workbook

from core.file_scanner import scan_json_files


MANUAL_LABEL_SEPARATORS = (',', '，', ';', '；', '\n', '\r')


def load_ordered_labels(file_path: str) -> tuple[list[str] | None, str | None]:
    """按原始顺序加载标签列表。

    参数:
        file_path: 标签文件路径

    返回:
        (标签列表, 错误信息) - 成功时 error 为 None
    """
    ext = os.path.splitext(file_path)[1].lower()

    try:
        if ext == '.csv':
            labels = _load_ordered_csv(file_path)
        elif ext == '.txt':
            labels = _load_ordered_txt(file_path)
        elif ext in ('.xlsx', '.xls'):
            labels = _load_ordered_excel(file_path)
        else:
            return None, f"不支持的文件格式: {ext}，仅支持 csv/txt/xlsx/xls"
    except Exception as e:
        return None, f"文件读取失败: {str(e)}"

    if not labels:
        return None, "标签文件中未找到有效的标签值"

    return labels, None


def parse_manual_ordered_labels(text: str) -> tuple[list[str] | None, str | None]:
    """解析手动输入的标签文本并保留首次出现顺序。"""
    normalized_text = text
    for separator in MANUAL_LABEL_SEPARATORS:
        normalized_text = normalized_text.replace(separator, '\n')

    labels = [item.strip() for item in normalized_text.splitlines() if item.strip()]
    labels = _deduplicate_labels(labels)
    if not labels:
        return None, "手动输入中未找到有效的标签值"

    return labels, None


def _deduplicate_labels(labels: list[str]) -> list[str]:
    """去重并保留首次出现顺序。"""
    seen = set()
    ordered_labels = []

    for label in labels:
        if label in seen:
            continue
        seen.add(label)
        ordered_labels.append(label)

    return ordered_labels


def _load_ordered_csv(file_path: str) -> list[str]:
    """加载 CSV 标签文件并保留顺序。"""
    labels = []
    with open(file_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if header is None or header[0].strip().lower() != 'label':
            raise ValueError("CSV 文件第一列列名必须为 'label'")

        for row in reader:
            if row and row[0].strip():
                labels.append(row[0].strip())

    return _deduplicate_labels(labels)


def _load_ordered_txt(file_path: str) -> list[str]:
    """加载 TXT 标签文件并保留顺序。"""
    labels = []
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            label = line.strip()
            if label:
                labels.append(label)

    return _deduplicate_labels(labels)


def _load_ordered_excel(file_path: str) -> list[str]:
    """加载 Excel 标签文件并保留顺序。"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("需要安装 openpyxl 库来读取 Excel 文件")

    labels = []
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    header = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = row[0]
            if str(header).strip().lower() != 'label':
                raise ValueError("Excel 文件第一列列名必须为 'label'")
            continue

        if row[0] is not None and str(row[0]).strip():
            labels.append(str(row[0]).strip())

    wb.close()
    return _deduplicate_labels(labels)


def run_label_counter(target_dir: str, ordered_labels: list[str]) -> tuple[str | None, str | None, dict]:
    """执行标签出现次数统计。

    参数:
        target_dir: 要统计的文件夹路径
        ordered_labels: 标签列表，保留原始顺序

    返回:
        (output_path, error, stats)
    """
    if not os.path.isdir(target_dir):
        return None, "目标文件夹不存在或不是有效目录", {}

    if not ordered_labels:
        return None, "标签列表不能为空", {}

    exclude_dirs = ["重叠检查结果", "抽样结果", "照片检查+"]
    all_json_files = list(scan_json_files(target_dir, exclude_dirs=exclude_dirs))

    total_files = len(all_json_files)
    if total_files == 0:
        return None, "目标文件夹内未找到 JSON 文件", {}

    label_set = set(ordered_labels)
    report_rows = []
    error_list = []

    for file_path in all_json_files:
        rel_path = os.path.relpath(file_path, target_dir)
        row_counts = {label: 0 for label in ordered_labels}

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            for shape in data.get('shapes', []):
                label = str(shape.get('label', '')).strip()
                if label in label_set:
                    row_counts[label] += 1

            report_rows.append((rel_path, row_counts))
        except Exception as e:
            error_list.append([rel_path, f"文件读取/解析失败: {str(e)}"])

    wb = Workbook()
    ws = wb.active
    ws.title = "统计报告"
    ws.append(["序号", "文件路径", *ordered_labels])

    for index, (rel_path, row_counts) in enumerate(report_rows, 1):
        ws.append([index, rel_path, *[row_counts[label] for label in ordered_labels]])

    if error_list:
        error_ws = wb.create_sheet(title="异常文件")
        error_ws.append(["序号", "文件路径", "错误信息"])
        for index, (rel_path, error_message) in enumerate(error_list, 1):
            error_ws.append([index, rel_path, error_message])

    output_path = os.path.join(target_dir, "标签出现次数统计报告.xlsx")
    wb.save(output_path)

    stats = {
        'total_files': total_files,
        'success_files': len(report_rows),
        'error_files': len(error_list),
        'label_count': len(ordered_labels),
    }

    return output_path, None, stats
