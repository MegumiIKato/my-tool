"""
标签校验工具 (Label Validator)

功能说明：
- 递归扫描选定目录中的所有 Labelme JSON 文件
- 检查每个 JSON 文件中的 shapes.label 值是否在用户提供的标签列表中
- 生成 xlsx 格式的检查报告，列出所有无效标签及其所在文件

使用场景：
- 标注数据质量检查
- 批量验证标签合法性
- 发现并修复错误的标签名称
"""

import os
import json
import csv
from openpyxl import Workbook

from core.file_scanner import scan_json_files


MANUAL_LABEL_SEPARATORS = (',', '，', ';', '；', '\n', '\r')


def load_label_dict(file_path: str) -> tuple[set[str] | None, str | None]:
    """加载标签列表文件
    
    支持格式: csv, txt, xlsx, xls
    第一列列名必须为 'label'
    
    参数:
        file_path: 标签文件路径
    
    返回:
        (标签集合, 错误信息) - 成功时 error 为 None
    """
    ext = os.path.splitext(file_path)[1].lower()
    
    labels = set()
    
    try:
        if ext == '.csv':
            labels = _load_csv(file_path)
        elif ext == '.txt':
            labels = _load_txt(file_path)
        elif ext in ('.xlsx', '.xls'):
            labels = _load_excel(file_path)
        else:
            return None, f"不支持的文件格式: {ext}，仅支持 csv/txt/xlsx/xls"
    except Exception as e:
        return None, f"文件读取失败: {str(e)}"
    
    if not labels:
        return None, "标签文件中未找到有效的标签值"
    
    return labels, None


def parse_manual_labels(text: str) -> tuple[set[str] | None, str | None]:
    """解析手动输入的标签文本。"""
    normalized_text = text
    for separator in MANUAL_LABEL_SEPARATORS:
        normalized_text = normalized_text.replace(separator, '\n')

    labels = {item.strip() for item in normalized_text.splitlines() if item.strip()}
    if not labels:
        return None, "手动输入中未找到有效的标签值"

    return labels, None


def _load_csv(file_path: str) -> set[str]:
    """加载 CSV 文件"""
    labels = set()
    with open(file_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if header is None or header[0].strip().lower() != 'label':
            raise ValueError("CSV 文件第一列列名必须为 'label'")
        for row in reader:
            if row and row[0].strip():
                labels.add(row[0].strip())
    return labels


def _load_txt(file_path: str) -> set[str]:
    """加载 TXT 文件，每行一个标签"""
    labels = set()
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line:
                labels.add(line)
    return labels


def _load_excel(file_path: str) -> set[str]:
    """加载 Excel 文件"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("需要安装 openpyxl 库来读取 Excel 文件")
    
    labels = set()
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    header = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = row[0]
            if str(header).strip().lower() != 'label':
                raise ValueError("Excel 文件第一列列名必须为 'label'")
            continue
        if row[0] is not None and str(row[0]).strip():
            labels.add(str(row[0]).strip())
    
    wb.close()
    return labels


def run_validator(target_dir: str, valid_labels: set[str]) -> tuple[str | None, str | None, dict]:
    """执行标签校验逻辑
    
    参数:
        target_dir: 要检查的文件夹路径
        valid_labels: 有效标签集合
    
    返回:
        (output_path, error, stats)
        - output_path: 报告文件路径，成功时返回
        - error: 错误信息，无错误返回 None
        - stats: {'total_files': int, 'error_count': int, 'valid_count': int}
    """
    if not os.path.isdir(target_dir):
        return None, "目标文件夹不存在或不是有效目录", {}

    if not valid_labels:
        return None, "标签列表不能为空", {}
    
    exclude_dirs = ["重叠检查结果", "抽样结果", "照片检查+"]
    all_json_files = list(scan_json_files(target_dir, exclude_dirs=exclude_dirs))

    total_files = len(all_json_files)
    if total_files == 0:
        return None, "目标文件夹内未找到 JSON 文件", {}

    error_list = []
    error_files = set()
    
    for file_path in all_json_files:
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                shapes = data.get('shapes', [])
                
                file_errors = set()
                for shape in shapes:
                    label = str(shape.get('label', "")).strip()
                    if label and label not in valid_labels:
                        file_errors.add(label)
                
                for err_lab in file_errors:
                    rel_path = os.path.relpath(file_path, target_dir)
                    error_list.append([rel_path, err_lab, "不在标签列表中"])
                    error_files.add(rel_path)
                    
        except Exception as e:
            rel_path = os.path.relpath(file_path, target_dir)
            error_list.append([rel_path, "", f"文件读取失败: {str(e)}"])
            error_files.add(rel_path)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "检查报告"
    ws.append(["序号", "文件路径", "无效标签", "错误类型"])
    
    for idx, (file_path, label, error_type) in enumerate(error_list, 1):
        ws.append([idx, file_path, label, error_type])
    
    if not error_list:
        ws.append([1, "-", "-", "未发现错误"])
    
    output_path = os.path.join(target_dir, "标签校验报告.xlsx")
    wb.save(output_path)
    
    stats = {
        'total_files': total_files,
        'error_count': len(error_files),
        'error_item_count': len(error_list),
        'valid_count': total_files - len(error_files)
    }
    
    return output_path, None, stats


def export_template(output_dir: str, file_type: str = 'csv') -> str:
    """导出空白模板文件
    
    参数:
        output_dir: 输出目录
        file_type: 模板类型 ('csv' 或 'xlsx')
    
    返回:
        模板文件路径
    """
    if file_type == 'csv':
        template_path = os.path.join(output_dir, "标签列表示例模板.csv")
        with open(template_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['label'])
    else:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "标签列表"
        ws.append(['label'])
        template_path = os.path.join(output_dir, "标签列表示例模板.xlsx")
        wb.save(template_path)
    
    return template_path


if __name__ == "__main__":
    import tkinter as tk
    from tkinter import filedialog, messagebox
    
    root = tk.Tk()
    root.withdraw()
    
    target_dir = filedialog.askdirectory(title="选择 Labelme JSON 所在文件夹")
    if not target_dir:
        print("未选择文件夹")
        exit()
    
    dict_file = filedialog.askopenfilename(
        title="选择标签文件",
        filetypes=[("标签文件", "*.csv *.txt *.xlsx *.xls")]
    )
    if not dict_file:
        print("未选择标签文件")
        exit()
    
    valid_labels, load_error = load_label_dict(dict_file)
    if load_error:
        messagebox.showerror("错误", load_error)
        exit()

    output_path, error, stats = run_validator(target_dir, valid_labels)
    
    if error:
        messagebox.showerror("错误", error)
    else:
        messagebox.showinfo(
            "完成",
            f"检查完成！\n\n总文件数: {stats['total_files']}\n有效文件: {stats['valid_count']}\n存在问题: {stats['error_count']}\n\n报告已保存至: {output_path}"
        )
